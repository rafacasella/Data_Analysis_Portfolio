#Percorrer Toda a base de dados. E para cada item:
    #verificar se o mesmo existe numa aba, se não existir criar aquela aba
    #copiar os valores daquela linha e colocar na aba correspondente

from openpyxl import load_workbook
from copy import copy #Copia os estilos de um objeto em python e cola em outro objeto de python (nesse caso consegue copiar as formatações das celulas)

def criar_aba(bairro, arquivo_base, estilo_cabecalho): #Criar a função criar aba. Está usando o arquivo_base ao invés de arquivo_bairros, para não usar variável global
    if bairro not in arquivo_base.sheetnames: #Se a aba não existe vai ser criada
        arquivo_base.create_sheet(bairro)
        nova_aba = arquivo_base[bairro]
        nova_aba['A1'].value = 'Data de Nascimento' #Criando cabeçalho nas abas novas
        nova_aba['B1'].value = 'Pessoa' #Colocando de forma manual pois é fixo
        nova_aba['C1'].value = 'Bairro'
        nova_aba['A1']._style = estilo_cabecalho #atribuindo a formatação criada
        nova_aba['B1']._style = estilo_cabecalho
        nova_aba['C1']._style = estilo_cabecalho

def transferir_informacoes_aba(aba_origem, aba_destino, linha_origem): #criar a função para transferir os dados
    linha_destino = aba_destino.max_row + 1 #ecrever na última linha após os dado anterior
    for coluna in range(1, 4): #Sempre colocar uma coluna a mais (nesse caso são 3), pois a última do range é excludente
        celula_origem = aba_origem.cell(row=linha_origem, column=coluna) #célula de origem está na aba de origem
        celula_destino = aba_destino.cell(row=linha_destino, column=coluna)
        celula_destino.value = celula_origem.value #passando efetivamente o valor da celula origem para destino
        celula_destino._style = copy(celula_origem._style)  #Copia o estilo, ou seja, formatação das células

arquivo_bairros = load_workbook('Bairros.xlsx')

print(arquivo_bairros.sheetnames) #Visualiza as abas que arquivo tem. É a aba que será usada como principal

aba_basedados = arquivo_bairros['Base de Dados'] #Garantir que a aba ativa será sempre a principal (base de dados)

ultima_linha = len(aba_basedados['A']) #Forma para garantir a quantidade de linhas máxima da planilha, (pego pela coluna A)
print(ultima_linha)

estilo_cabecalho = copy(aba_basedados['A1']._style) #formatação do cabeçalho, está fora do OR, pq só vai usar 1 vez

for linha in range(2, ultima_linha + 1): #Um For para percorrer todas as linhas até a última (começa da 2, pq 1 é cabeçalho)
    bairro = aba_basedados.cell(row=linha, column=3).value # armazena na variável bairro, da aba principal, todas as linhas da coluna 3
    #onde se encontra o dado que será analisado. bairro = aba_basedados[f"C{linha}"].value , outra maneira de ser feito.
    if not bairro:
        break #Se não encontrar nenhum dado na coluna bairro (célula em branco), pra o FOR
    #Caso encontre o bairro, deve-se criar uma aba pro bairro e transferir as informações para ela
    #Para criar a aba, é necessária uma função com o parâmetro bairro
    criar_aba(bairro, arquivo_bairros, estilo_cabecalho)

    #Para transferir as informações é necessária passar os parâmetros aba origem, aba destino e linha origem
    aba_destino = arquivo_bairros[bairro]
    transferir_informacoes_aba(aba_basedados, aba_destino, linha)

arquivo_bairros.save('Bairros2.xlsx')