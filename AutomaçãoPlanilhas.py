#Automação de planilhas (Comandos Básicos)
from openpyxl import load_workbook #para não importar a biblioteca inteira porque é enorme

planilha = load_workbook('Vendas.xlsx') #Caso a planilha não esteja na mesma pasta, tem que passar o caminho inteiro

#Ver as abas da planilha
print(planilha.sheetnames)

#Pegar a Aba Ativa (a que está selecionada no momento)
aba_atual = planilha.active
print(aba_atual)

#Selecionar uma aba específica
aba_especifica = planilha['Plan1']
print(aba_especifica)

#Selecionar uma célula específica
valor_B1 = aba_especifica['B1'].value
print(valor_B1)
valor_B2 = aba_especifica.cell(row = 2, column = 2).value
print(valor_B2)

#Editar as células
aba_atual.cell(row = 1, column = 2).value = "Datas"
#Sempre é necessário salvar após editar
planilha.save('Vendas2.xlsx') #Se salvar com o mesmo nome substitui o arquivo

#Pegar a última linha de uma coluna, ou seja, percorrer até o final
print(aba_atual.max_row) #Pode dar valor errado, errando para mais, pois ele conta ultimo preenchido
print(aba_atual.max_column)

print(len(aba_atual['A'])) #Método mais eficaz
